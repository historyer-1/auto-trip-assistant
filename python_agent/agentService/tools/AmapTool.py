"""高德地图查询工具。"""

from __future__ import annotations

from functools import lru_cache
from pathlib import Path
from typing import Any

import httpx

from agentService.entity.api_keys import AMAP_WEB_API_KEY
from langchain.tools import tool
from openpyxl import load_workbook


@lru_cache(maxsize=1)
def _load_adcode_mapping() -> dict[str, str]:
	"""从本地 AMap_adcode_citycode.xlsx 加载城市与 adcode 对照表。

	参数:
		无。

	返回值:
		dict[str, str]: key 为城市名，value 为 adcode。
	"""
	mapping: dict[str, str] = {}
	xlsx_path = Path(__file__).with_name("AMap_adcode_citycode.xlsx")

	if not xlsx_path.exists():
		return mapping

	workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
	try:
		sheet = workbook.active
		if sheet is None:
			return mapping
		headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
		if not headers:
			return mapping

		name_index = -1
		adcode_index = -1
		for idx, header in enumerate(headers):
			header_text = str(header or "").strip().lower()
			if header_text == "中文名" and name_index == -1:
				name_index = idx #城市对应列数
			if  header_text == "adcode" and adcode_index == -1:
				adcode_index = idx #adcode对应列数

		if name_index == -1 or adcode_index == -1:
			return mapping

		for row in sheet.iter_rows(min_row=2, values_only=True):
			if not row or len(row) <= max(name_index, adcode_index):
				continue

			city_name = str(row[name_index] or "").strip()
			adcode = str(row[adcode_index] or "").strip()
			if not city_name or not adcode.isdigit():
				continue

			mapping[city_name] = adcode
			if city_name.endswith("市") or city_name.endswith("区") or city_name.endswith("县"):
				mapping[city_name[:-1]] = adcode	
	finally:
		workbook.close()

	return mapping


def _resolve_adcode(city: str) -> str:
	"""将城市名称或 adcode 转为可直接请求高德天气接口的 adcode。

	参数:
		city: 城市名称（如“北京”）或 6 位 adcode。

	返回值:
		str: adcode；无法解析时返回空字符串。
	"""
	city_text = (city or "").strip()

	mapping = _load_adcode_mapping()
	if city_text in mapping:
		return mapping[city_text]

	with_suffix = f"{city_text}市"
	if with_suffix in mapping:
		return mapping[with_suffix]

	return ""

@tool
def amap_search(keywords: str, region: str, page_size: int = 5) -> str:
	"""按关键字和城市区划查询高德地点信息。

	参数:
		keywords: 查询关键字，例如“北京大学”，只接受一个关键词。
		region: 搜索城市区划，例如“北京市”，只接受一个城市。
		page_size: 返回条数，默认 5，范围 1 到 6。

	返回值:
		str: 统一格式的字符串结果；成功时返回地点信息，失败时返回错误信息。
	"""

	# 先做最小必要的参数校验，避免把明显错误请求发给高德接口。
	if not keywords or not keywords.strip():
		return "status=0; error=keywords 不能为空"

	if not region or not region.strip():
		return "status=0; error=region 不能为空"

	# page_size 只允许 1~6，避免一次返回过多结果导致上下文膨胀。
	page_size = max(1, min(int(page_size), 6))

	url = "https://restapi.amap.com/v5/place/text"
	params = {
		"keywords": keywords.strip(),
		"region": region.strip(),
		"page_size": page_size,
		# 只返回本次需要的字段：状态、商业信息、图片信息。
		"show_fields": "business,photos",
		"key": AMAP_WEB_API_KEY,
	}

	try:
		# 直接调用高德文本检索接口，结果由 show_fields 控制返回字段。
		response = httpx.get(url, params=params, timeout=10.0)
		response.raise_for_status()
		data: dict[str, Any] = response.json()
	except Exception as exc:
		return f"status=0; error=高德API请求失败: {exc}"

	# 高德接口本身会返回 status 字段，1 表示成功，0 表示失败。
	status = str(data.get("status", "0"))
	if status != "1":
		return f"status=0; info={data.get('info', '查询失败')}; infocode={data.get('infocode', '')}"

	# 结果字符串保持简洁，方便 Agent 直接阅读和继续总结。
	pois = data.get("pois") or []
	if not isinstance(pois, list) or not pois:
		return "status=1; data=未找到匹配结果"

	lines = ["status=1"]
	for index, poi in enumerate(pois[:page_size], start=1):
		if not isinstance(poi, dict):
			continue

		name = poi.get("name", "")
		address = poi.get("address", "")
		business = poi.get("business", {})
		photos = poi.get("photos", [])

		# 将商业信息和图片信息尽量保持原样输出，减少无谓的二次加工。
		lines.append(
			f"{index}. name={name}; address={address}; business={business}; photos={photos}"
		)

	if len(lines) == 1:
		return "status=1; data=未找到可解析的地点信息"

	return "\n".join(lines)


@tool
def amap_weather_search(city: str) -> str:
	"""按城市查询高德天气预报。

	参数:
		city: 城市名称，只接受一个城市。

	返回值:
		str: 统一格式字符串；成功返回天气预报，失败返回错误信息。
	"""
	if not city or not city.strip():
		return "status=0; error=city 不能为空"

	adcode = _resolve_adcode(city)
	if not adcode:
		return ("status=0; error=无法解析城市adcode")

	url = "https://restapi.amap.com/v3/weather/weatherInfo"
	params = {
		"key": AMAP_WEB_API_KEY,
		"city": adcode,
		# 使用 all 获取未来多天预报，便于行程规划按天拆分。
		"extensions": "all",
	}

	try:
		response = httpx.get(url, params=params, timeout=10.0)
		response.raise_for_status()
		data: dict[str, Any] = response.json()
	except Exception as exc:
		return f"status=0; error=高德天气API请求失败: {exc}"

	status = str(data.get("status", "0"))
	if status != "1":
		return f"status=0; info={data.get('info', '查询失败')}; infocode={data.get('infocode', '')}"

	forecasts = data.get("forecasts") or []
	if not isinstance(forecasts, list) or not forecasts:
		return "status=1; data=未找到天气预报数据"

	forecast = forecasts[0] if isinstance(forecasts[0], dict) else {}
	city_name = forecast.get("city", "")
	city_adcode = forecast.get("adcode", adcode)
	casts = forecast.get("casts") or []
	if not isinstance(casts, list) or not casts:
		return "status=1; data=未找到逐日天气信息"

	lines = [f"status=1; city={city_name}; adcode={city_adcode}"]
	for index, cast in enumerate(casts, start=1):
		if not isinstance(cast, dict):
			continue

		date = cast.get("date", "")
		day_weather = cast.get("dayweather", "")
		night_weather = cast.get("nightweather", "")
		day_temp = cast.get("daytemp", "")
		night_temp = cast.get("nighttemp", "")
		day_wind = cast.get("daywind", "")
		night_wind = cast.get("nightwind", "")
		day_power = cast.get("daypower", "")
		night_power = cast.get("nightpower", "")

		# 同一行输出白天/夜间信息，方便模型一次解析成结构化天气列表。
		lines.append(
			f"{index}. date={date}; day_weather={day_weather}; night_weather={night_weather}; "
			f"day_temp={day_temp}; night_temp={night_temp}; "
			f"day_wind={day_wind}; night_wind={night_wind}; "
			f"day_power={day_power}; night_power={night_power}"
		)

	if len(lines) == 1:
		return "status=1; data=未找到可解析的天气信息"

	return "\n".join(lines)

